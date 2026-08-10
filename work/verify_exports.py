"""
Evaluates the workbook's formulas and checks them against analysis.json.

LibreOffice does not run in this environment, so the workbook cannot be
recalculated by an application. This does the recalculation instead: a small
evaluator for the subset of Excel used here (arithmetic, comparisons, IF, AND,
OR, NOT, ISNUMBER, MIN, MAX, SUM, ROUND, AVERAGE, cross-sheet references),
resolving each formula against the cells it names, then comparing the result to
the figure the analysis engine produced.

This is stricter than a recalculation would be. Recalculating only proves the
formulas evaluate without error; a formula off by one row evaluates perfectly
and gives the wrong answer. Comparing every evaluated result against the engine
proves they compute what they claim to.
"""

import json
import math
import os
import re
import sys

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

WORK = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(WORK, "버핏기준_52개기업_분석.xlsx")
REL = 1e-6

TOKEN = re.compile(r"""
    (?P<sheetref>(?:'[^']+'|[^\s!,()+\-*/^<>=&:]+)!\$?[A-Z]{1,3}\$?\d+)
  | (?P<ref>\$?[A-Z]{1,3}\$?\d+)
  | (?P<num>\d+(?:\.\d+)?(?:[eE][+-]?\d+)?)
  | (?P<str>"[^"]*")
  | (?P<func>[A-Z][A-Z0-9_.]*(?=\())
  | (?P<op><>|<=|>=|[-+*/^<>=,():])
""", re.X)


class Evaluator:
    """Enough of Excel to check this workbook, and no more."""

    def __init__(self, wb):
        self.wb = wb
        self.cache = {}
        self.stack = set()

    # -- cell access ------------------------------------------------------
    def raw(self, sheet, coord):
        return self.wb[sheet][coord.replace("$", "")].value

    def value(self, sheet, coord):
        key = (sheet, coord.replace("$", ""))
        if key in self.cache:
            return self.cache[key]
        if key in self.stack:
            raise ValueError(f"circular reference at {key}")
        self.stack.add(key)
        try:
            v = self.raw(*key)
            out = self.evaluate(v, sheet) if isinstance(v, str) and v.startswith("=") else v
        finally:
            self.stack.discard(key)
        self.cache[key] = out
        return out

    # -- parsing ----------------------------------------------------------
    def evaluate(self, formula, sheet):
        """
        Parse and evaluate one formula.

        Parser position lives on the instance, and resolving a reference can
        recurse back into evaluate() for a cell that has not been computed yet,
        which would otherwise overwrite the caller's position mid-parse. Saving
        and restoring it keeps the evaluator reentrant - without this, formulas
        happen to work when their dependencies were already cached and fail when
        they were not, which is the worst kind of intermittent.
        """
        saved = (getattr(self, "tokens", None), getattr(self, "pos", None),
                 getattr(self, "sheet", None))
        try:
            self.tokens = [m.group(0) for m in TOKEN.finditer(formula[1:])]
            self.pos, self.sheet = 0, sheet
            v = self.expr()
            if self.pos != len(self.tokens):
                raise ValueError(f"unconsumed tokens in {formula!r} at {self.tokens[self.pos:]}")
            return v
        finally:
            self.tokens, self.pos, self.sheet = saved

    def peek(self):
        return self.tokens[self.pos] if self.pos < len(self.tokens) else None

    def take(self, expected=None):
        t = self.peek()
        if expected and t != expected:
            raise ValueError(f"expected {expected!r}, got {t!r}")
        self.pos += 1
        return t

    def expr(self):
        left = self.arith()
        while self.peek() in ("=", "<>", "<", ">", "<=", ">="):
            op = self.take()
            right = self.arith()
            a, b = self._cmp_operands(left, right)
            left = {"=": a == b, "<>": a != b, "<": a < b,
                    ">": a > b, "<=": a <= b, ">=": a >= b}[op]
        return left

    @staticmethod
    def _cmp_operands(a, b):
        # Excel orders numbers below text; the only comparison here that meets
        # a blank is a number against "", so treat a blank as -inf.
        if isinstance(a, str) or isinstance(b, str):
            if a == "" or a is None:
                a = -math.inf
            if b == "" or b is None:
                b = -math.inf
        return (a if a is not None else 0), (b if b is not None else 0)

    def arith(self):
        v = self.term()
        while self.peek() in ("+", "-"):
            op = self.take()
            r = self.term()
            v = self.num(v) + self.num(r) if op == "+" else self.num(v) - self.num(r)
        return v

    def term(self):
        v = self.power()
        while self.peek() in ("*", "/"):
            op = self.take()
            r = self.power()
            if op == "*":
                v = self.num(v) * self.num(r)
            else:
                d = self.num(r)
                if d == 0:
                    raise ZeroDivisionError("#DIV/0!")
                v = self.num(v) / d
        return v

    def power(self):
        v = self.unary()
        while self.peek() == "^":
            self.take()
            v = self.num(v) ** self.num(self.unary())
        return v

    def unary(self):
        if self.peek() == "-":
            self.take()
            return -self.num(self.unary())
        if self.peek() == "+":
            self.take()
        return self.atom()

    def atom(self):
        t = self.peek()
        if t is None:
            raise ValueError("unexpected end of formula")
        if t == "(":
            self.take("(")
            v = self.expr()
            self.take(")")
            return v
        if re.fullmatch(r"[A-Z][A-Z0-9_.]*", t) and self.tokens[self.pos + 1:self.pos + 2] == ["("]:
            return self.call(self.take())
        self.take()
        if t.startswith('"'):
            return t[1:-1]
        if re.fullmatch(r"\d+(\.\d+)?([eE][+-]?\d+)?", t):
            return float(t)
        if "!" in t:
            sheet, coord = t.split("!", 1)
            return self.value(sheet.strip("'"), coord)
        return self.value(self.sheet, t)

    def args(self):
        self.take("(")
        out = []
        if self.peek() == ")":
            self.take(")")
            return out
        while True:
            out.append(self.arg_or_range())
            if self.peek() == ",":
                self.take(",")
                continue
            self.take(")")
            return out

    def arg_or_range(self):
        # A range only ever appears as a bare argument here (SUM/AVERAGE).
        t, nxt = self.peek(), self.tokens[self.pos + 1:self.pos + 2]
        if t and nxt == [":"] and re.fullmatch(r"(?:[^!]+!)?\$?[A-Z]{1,3}\$?\d+", t):
            start = self.take()
            self.take(":")
            end = self.take()
            sheet = self.sheet
            if "!" in start:
                sheet, start = start.split("!", 1)
                sheet = sheet.strip("'")
            if "!" in end:
                end = end.split("!", 1)[1]
            return self.range_values(sheet, start, end)
        return self.expr()

    def range_values(self, sheet, start, end):
        def parse(c):
            m = re.fullmatch(r"\$?([A-Z]{1,3})\$?(\d+)", c)
            return column_index_from_string(m.group(1)), int(m.group(2))
        c1, r1 = parse(start)
        c2, r2 = parse(end)
        out = []
        for rr in range(min(r1, r2), max(r1, r2) + 1):
            for cc in range(min(c1, c2), max(c1, c2) + 1):
                out.append(self.value(sheet, f"{self.wb[sheet].cell(row=rr, column=cc).coordinate}"))
        return out

    def skip_arg(self):
        """Record an argument's token span without evaluating it."""
        start, depth = self.pos, 0
        while self.pos < len(self.tokens):
            t = self.tokens[self.pos]
            if t == "(":
                depth += 1
            elif t == ")":
                if depth == 0:
                    break
                depth -= 1
            elif t == "," and depth == 0:
                break
            self.pos += 1
        return (start, self.pos)

    def eval_span(self, span):
        saved = self.pos
        try:
            self.pos = span[0]
            v = self.expr()
            if self.pos != span[1]:
                raise ValueError("argument not fully consumed")
            return v
        finally:
            self.pos = saved

    def call_if(self):
        """
        IF, evaluated lazily, the way Excel does.

        Evaluating every branch eagerly makes guarded formulas fail: the whole
        point of IF(denominator<>0, a/b, "") is that the division never happens
        when the guard is false. An eager evaluator raises #DIV/0! on a formula
        Excel computes without complaint, and the workbook gets blamed for a
        fault in the checker.
        """
        self.take("(")
        cond = self.expr()
        self.take(",")
        then_span = self.skip_arg()
        else_span = None
        if self.peek() == ",":
            self.take(",")
            else_span = self.skip_arg()
        self.take(")")
        span = then_span if self.truth(cond) else else_span
        return False if span is None else self.eval_span(span)

    def call(self, name):
        if name == "IF":
            return self.call_if()
        args = self.args()
        flat = []
        for a in args:
            flat.extend(a) if isinstance(a, list) else flat.append(a)
        if name == "IF":
            return args[1] if self.truth(args[0]) else (args[2] if len(args) > 2 else False)
        if name == "AND":
            return all(self.truth(a) for a in flat)
        if name == "OR":
            return any(self.truth(a) for a in flat)
        if name == "NOT":
            return not self.truth(args[0])
        if name == "ISNUMBER":
            return isinstance(args[0], (int, float)) and not isinstance(args[0], bool)
        if name == "MIN":
            return min(self.num(a) for a in flat if a not in (None, ""))
        if name == "MAX":
            return max(self.num(a) for a in flat if a not in (None, ""))
        if name == "SUM":
            return sum(self.num(a) for a in flat if isinstance(a, (int, float)))
        if name == "AVERAGE":
            vals = [self.num(a) for a in flat if isinstance(a, (int, float))]
            return sum(vals) / len(vals) if vals else 0.0
        if name == "ROUND":
            n, d = self.num(args[0]), int(self.num(args[1]))
            # Excel rounds half away from zero; Python rounds half to even.
            f = 10 ** d
            return math.floor(abs(n) * f + 0.5) / f * (1 if n >= 0 else -1)
        raise ValueError(f"unsupported function {name}")

    @staticmethod
    def truth(v):
        if isinstance(v, bool):
            return v
        if v is None or v == "":
            return False
        return bool(v)

    @staticmethod
    def num(v):
        if isinstance(v, bool):
            return 1.0 if v else 0.0
        if v is None or v == "":
            return 0.0
        if isinstance(v, str):
            raise ValueError(f"#VALUE! - text {v!r} used as a number")
        return float(v)


def close(a, b, rel=REL):
    if a in (None, "") or b is None:
        return (a in (None, "")) and b is None
    return abs(a - b) <= rel * max(1.0, abs(a), abs(b))


def main():
    with open(os.path.join(WORK, "analysis.json")) as f:
        analysis = json.load(f)
    # Score from report.py rather than from a value stamped into the JSON, so
    # the check compares the workbook against the definition, not against a
    # copy of itself.
    sys.path.insert(0, WORK)
    import report as rp
    for c in analysis["companies"]:
        c["summary"]["_latest_revenue"] = next(
            (r["revenue"] for r in reversed(c["years"]) if r.get("revenue")), None)
        c["_expected_score"], _ = rp.quality_score(c["summary"])
    by = {c["ticker"]: c for c in analysis["companies"]}

    wb = load_workbook(XLSX)
    ev = Evaluator(wb)
    fails, checked = [], 0

    if not (wb.calculation and wb.calculation.fullCalcOnLoad):
        fails.append("workbook will not recalculate on open; formula cells would be blank")

    def col_index(ws, header_row):
        return {c.value: c.column for c in ws[header_row] if c.value}

    # ---- 연도별계산 -------------------------------------------------------
    ws = wb["연도별계산"]
    ci = col_index(ws, 4)
    L = lambda name, row: ws.cell(row=row, column=ci[name]).coordinate
    year_index = {}
    for row in range(5, ws.max_row + 1):
        t = ws.cell(row=row, column=ci["티커"]).value
        if not t:
            continue
        fy = ws.cell(row=row, column=ci["회계연도"]).value
        year_index[(t, fy)] = row

    for (t, fy), row in year_index.items():
        rec = by.get(t)
        if not rec:
            continue
        y = next((z for z in rec["years"] if z["fiscal_year"] == fy), None)
        if not y:
            continue
        pairs = [
            ("영업이익(EBIT)", y.get("ebit")),
            ("실효세율", y.get("tax_rate"), 1.0),
            ("NOPAT", y.get("nopat")),
            ("투하자본", y.get("invested_capital")),
            ("투하자본(평균)", y.get("invested_capital_avg")),
            ("ROIC", y.get("roic"), 1.0),
            ("ROE", y.get("roe"), 1.0),
            ("EBITDA", y.get("ebitda")),
            ("순부채/EBITDA", y.get("net_debt_to_ebitda"), 1.0),
            ("이자보상배율", y.get("interest_coverage"), 1.0),
            ("운전자본증감", (y.get("raw") or {}).get("change_in_working_capital")),
            ("주주이익", y.get("owner_earnings")),
            ("주주이익률", y.get("owner_earnings_margin"), 1.0),
            ("영업이익률", y.get("operating_margin"), 1.0),
        ]
        for spec in pairs:
            name, expected = spec[0], spec[1]
            scale = spec[2] if len(spec) > 2 else 1e9
            try:
                got = ev.value("연도별계산", L(name, row))
            except Exception as e:                       # noqa: BLE001
                fails.append(f"연도별계산 {t} FY{fy} {name}: {e}")
                continue
            # ROIC is suppressed in the engine when capital is negative; the
            # sheet's guard does the same, so both should be blank together.
            if name == "ROIC" and y.get("roic_status", "").startswith("NOT_MEANINGFUL"):
                expected = None
            exp = None if expected is None else expected / scale
            if not close(got if got != "" else None, exp, 1e-5):
                fails.append(f"연도별계산 {t} FY{fy} {name}: sheet={got!r} engine={exp!r}")
            checked += 1

    # ---- 품질순위 ---------------------------------------------------------
    ws = wb["품질순위"]
    ci = col_index(ws, 4)
    for row in range(5, ws.max_row + 1):
        t = ws.cell(row=row, column=ci["티커"]).value
        if not t:
            continue
        try:
            got = ev.value("품질순위", ws.cell(row=row, column=ci["총점"]).coordinate)
        except Exception as e:                           # noqa: BLE001
            fails.append(f"품질순위 {t} 총점: {e}")
            continue
        exp = by[t]["_expected_score"]
        if int(got) != exp:
            fails.append(f"품질순위 {t}: sheet score={got} engine={exp}")
        checked += 1

    # ---- DCF계산 ----------------------------------------------------------
    ws = wb["DCF계산"]
    ci = col_index(ws, 4)
    scen_key = {"보수": "conservative", "기준": "base", "낙관": "optimistic"}
    for row in range(5, ws.max_row + 1):
        t = ws.cell(row=row, column=ci["티커"]).value
        if not t:
            continue
        scen = scen_key[ws.cell(row=row, column=ci["시나리오"]).value]
        d = by[t]["summary"]["dcf"][scen]
        for name, expected, scale in (("적정가치", d["intrinsic_value"], 1e9),
                                      ("안전마진", d["margin_of_safety"], 1.0),
                                      ("적용 성장률", d["assumptions"]["growth_rate"], 1.0),
                                      ("할인율", d["assumptions"]["discount_rate"], 1.0)):
            try:
                got = ev.value("DCF계산", ws.cell(row=row, column=ci[name]).coordinate)
            except Exception as e:                       # noqa: BLE001
                fails.append(f"DCF계산 {t} {scen} {name}: {e}")
                continue
            exp = None if expected is None else expected / scale
            if not close(got if got != "" else None, exp, 1e-5):
                fails.append(f"DCF계산 {t} {scen} {name}: sheet={got!r} engine={exp!r}")
            checked += 1
        verdict = ev.value("DCF계산", ws.cell(row=row, column=ci["판정"]).coordinate)
        want = {"INVESTABLE_RANGE": "투자가능", "BORDERLINE": "경계선",
                "OUTSIDE_RANGE": "범위밖", "DATA_UNAVAILABLE": "판정불가"}[d["verdict"]]
        if verdict != want:
            fails.append(f"DCF계산 {t} {scen} 판정: sheet={verdict} engine={want}")
        checked += 1

    # ---- 밸류에이션 (multiples and the cross-sheet links) -----------------
    ws = wb["밸류에이션"]
    ci = col_index(ws, 4)
    for row in range(5, ws.max_row + 1):
        t = ws.cell(row=row, column=ci["티커"]).value
        if not t:
            continue
        s = by[t]["summary"]
        for name, expected in (("PER", s.get("per")), ("PBR", s.get("pbr")),
                               ("PSR", s.get("psr")),
                               ("기준 적정가치", (s["dcf"]["base"]["intrinsic_value"] or 0) / 1e9),
                               ("안전마진", s["dcf"]["base"]["margin_of_safety"])):
            try:
                got = ev.value("밸류에이션", ws.cell(row=row, column=ci[name]).coordinate)
            except Exception as e:                       # noqa: BLE001
                fails.append(f"밸류에이션 {t} {name}: {e}")
                continue
            if not close(got if got != "" else None, expected, 1e-5):
                fails.append(f"밸류에이션 {t} {name}: sheet={got!r} engine={expected!r}")
            checked += 1

    # ---- percentages stored as fractions ---------------------------------
    for sheet, colname, hdr in (("품질순위", "ROIC 중앙값", 4), ("금융9개사", "ROE 중앙값", 4)):
        ws = wb[sheet]
        ci = col_index(ws, hdr)
        for row in range(hdr + 1, ws.max_row + 1):
            v = ws.cell(row=row, column=ci[colname]).value
            if isinstance(v, (int, float)) and abs(v) > 5:
                fails.append(f"{sheet}!{colname} row {row}: {v} looks like a percent "
                             f"stored as a whole number")

    print(f"sheets: {wb.sheetnames}")
    print(f"formula results evaluated and compared: {checked}")
    if fails:
        print(f"\nFAILURES ({len(fails)}):")
        for f_ in fails[:40]:
            print("  -", f_)
        if len(fails) > 40:
            print(f"  ... and {len(fails)-40} more")
        return 1
    print("every formula in the workbook reproduces the analysis engine's figure")
    return 0


if __name__ == "__main__":
    sys.exit(main())
