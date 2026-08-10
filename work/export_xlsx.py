"""
Excel export - the analysis as a workbook whose arithmetic is visible.

The point of this build is that a reader can click any derived cell and see the
calculation in the formula bar, then follow its references back to figures taken
straight from a 10-K. Nothing that can be computed inside the workbook is
imported as a constant.

Colour says where a number came from, following the usual model convention:
  blue   - a raw input, lifted from a filing or a quote, edit-able
  black  - a formula computed from other cells in the workbook
  green  - a formula that reaches into another sheet
  yellow - an assumption you may want to change

So a row of 연도별계산 reads left to right as: what the filing said, then every
ratio built from it. 기준표 holds every threshold and discount rate as its own
cell, so changing one moves the scores and valuations that depend on it.

Writes work/버핏기준_52개기업_분석.xlsx.
"""

import json
import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.properties import CalcProperties

WORK = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(WORK, "버핏기준_52개기업_분석.xlsx")

FONT = "Arial"
PCT = "0.0%"
PCT2 = "0.00%"
MULT = '0.0"x"'
USD_B = '$#,##0.0;($#,##0.0);-'
INT = "#,##0"
NUM = "#,##0.00"

BLUE = Font(name=FONT, size=10, color="0000FF")          # raw input
BLACK = Font(name=FONT, size=10)                          # formula, same sheet
GREEN = Font(name=FONT, size=10, color="008000")          # formula, other sheet
BOLD = Font(name=FONT, size=10, bold=True)
NOTE = Font(name=FONT, size=9, italic=True, color="595959")
TITLE_FONT = Font(name=FONT, bold=True, size=14, color="1F3864")
HEAD_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=9)
HEAD_FILL = PatternFill("solid", fgColor="1F3864")
RAW_FILL = PatternFill("solid", fgColor="DEEAF6")         # raw-input column band
YELLOW = PatternFill("solid", fgColor="FFFF00")
BORDER = Border(bottom=Side(style="thin", color="D9D9D9"))


def bn(v):
    return None if v is None else v / 1e9


def title_block(ws, title, subtitle=None):
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    if subtitle:
        ws["A2"] = subtitle
        ws["A2"].font = NOTE
        ws["A2"].alignment = Alignment(vertical="top")
    return 4 if subtitle else 3


def header(ws, row, labels, widths=None, raw_cols=()):
    for i, label in enumerate(labels, 1):
        c = ws.cell(row=row, column=i, value=label)
        c.fill = HEAD_FILL
        c.font = HEAD_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    if widths:
        for i, wd in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = wd
    ws.row_dimensions[row].height = 34
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def put(ws, row, col, value, fmt=None, font=BLACK, fill=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font
    c.border = BORDER
    if fmt:
        c.number_format = fmt
    if fill:
        c.fill = fill
    return c


# ===========================================================================
# 기준표 - every assumption and threshold, in its own cell
# ===========================================================================

ASSUMPTIONS = [
    ("세율", None, None, None),
    ("실효세율 하한", 0.05, PCT, "이보다 낮은 실효세율은 일회성 항목의 결과로 보고 하한을 적용"),
    ("실효세율 상한", 0.40, PCT, "이보다 높은 실효세율도 마찬가지로 상한을 적용"),
    ("기본 세율", 0.21, PCT, "세전이익이 없거나 음수여서 실효세율을 낼 수 없을 때 쓰는 미국 연방 법정세율"),
    ("품질 배점 기준", None, None, None),
    ("ROIC 25점 기준", 0.30, PCT, "ROIC 중앙값이 이 이상이면 25점"),
    ("ROIC 20점 기준", 0.20, PCT, None),
    ("ROIC 15점 기준", 0.15, PCT, None),
    ("ROIC 10점 기준", 0.10, PCT, None),
    ("신규ROIC 20점 기준", 0.20, PCT, "증분 ROIC가 이 이상이면 20점"),
    ("신규ROIC 15점 기준", 0.15, PCT, None),
    ("신규ROIC 10점 기준", 0.10, PCT, None),
    ("스프레드 15점 기준", 0.15, PCT, "ROIC−WACC가 이 이상이면 15점"),
    ("스프레드 12점 기준", 0.10, PCT, None),
    ("스프레드 8점 기준", 0.05, PCT, None),
    ("주주이익률 10점 기준", 0.20, PCT, "정규화 주주이익 ÷ 매출이 이 이상이면 10점"),
    ("주주이익률 7점 기준", 0.10, PCT, None),
    ("주주이익률 4점 기준", 0.05, PCT, None),
    ("순부채 5점 기준", 1.0, MULT, "순부채/EBITDA가 이 이하면 5점"),
    ("순부채 4점 기준", 2.0, MULT, None),
    ("순부채 2점 기준", 3.0, MULT, None),
    ("이자보상 5점 기준", 10.0, MULT, "이자보상배율이 이 이상이면 5점"),
    ("이자보상 3점 기준", 5.0, MULT, None),
    ("이자보상 1점 기준", 2.0, MULT, None),
    ("DCF 가정", None, None, None),
    ("예측기간(년)", 10, INT, "명시적으로 현금흐름을 추정하는 기간"),
    ("보수 할인율", 0.12, PCT, None),
    ("보수 영구성장률", 0.020, PCT, None),
    ("보수 성장률 배수", 0.5, NUM, "과거 주주이익 성장률에 곱하는 배수"),
    ("기준 할인율", 0.10, PCT, None),
    ("기준 영구성장률", 0.025, PCT, None),
    ("기준 성장률 배수", 1.0, NUM, None),
    ("낙관 할인율", 0.08, PCT, None),
    ("낙관 영구성장률", 0.030, PCT, None),
    ("낙관 성장률 배수", 1.25, NUM, None),
    ("성장률 상한(보수·기준)", 0.10, PCT, "10년을 이 속도보다 빠르게 복리성장한다고 보지 않음"),
    ("성장률 상한(낙관)", 0.15, PCT, None),
    ("안전마진 투자가능 기준", 0.30, PCT, "안전마진이 이 이상이면 투자 가능 구간"),
    ("매크로", None, None, None),
]


def sheet_assumptions(wb, analysis):
    ws = wb.create_sheet("기준표")
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 78
    r = title_block(ws, "기준표 — 가정과 판정 기준",
                    "노란 셀은 바꿀 수 있는 가정입니다. 다른 시트의 수식이 이 셀들을 참조하므로, "
                    "값을 바꾸면 점수와 적정가치가 함께 움직입니다.")
    header(ws, r, ["항목", "값", "설명"], [30, 14, 78])
    r += 1
    ref = {}
    for name, value, fmt, desc in ASSUMPTIONS:
        if value is None:
            put(ws, r, 1, name, font=BOLD)
            r += 1
            continue
        put(ws, r, 1, name)
        c = put(ws, r, 2, value, fmt, font=BLUE, fill=YELLOW)
        put(ws, r, 3, desc or "", font=NOTE)
        ref[name] = f"기준표!$B${r}"
        r += 1

    rf, erp = analysis["risk_free_rate"], analysis["equity_risk_premium"]
    for name, value, fmt, desc in (
        ("무위험수익률", rf["rate"], PCT2, f"{rf['series']} · {rf['as_of']} · {rf['source_url']}"),
        ("주식위험프리미엄", erp["erp"], PCT2,
         f"Damodaran (NYU Stern) · {erp['as_of']} · {erp['source_url']}"),
    ):
        put(ws, r, 1, name)
        put(ws, r, 2, value, fmt, font=BLUE, fill=YELLOW)
        put(ws, r, 3, desc, font=NOTE)
        ref[name] = f"기준표!$B${r}"
        r += 1
    return ref


# ===========================================================================
# 연도별계산 - raw filing inputs, then every ratio built from them
# ===========================================================================

# (header, width, kind, key/format). kind: "raw" | "text" | formula template
YEAR_COLS = [
    ("티커", 9, "text", "ticker"),
    ("기업", 22, "text", "company_name"),
    ("회계연도", 9, "text", "fiscal_year"),
    ("결산일", 11, "text", "period_end"),
    ("업종구분", 10, "text", "sector"),
    ("접수번호", 21, "text", "accession"),
    # ---- raw ----
    ("매출", 13, "raw", ("revenue", USD_B)),
    ("세전이익", 13, "raw", ("pretax_income", USD_B)),
    ("이자비용", 12, "raw", ("interest_expense", USD_B)),
    ("법인세비용", 12, "raw", ("income_tax_expense", USD_B)),
    ("영업이익(보고)", 13, "raw", ("operating_income_reported", USD_B)),
    ("순이익", 13, "raw", ("net_income", USD_B)),
    ("감가상각", 12, "raw", ("depreciation_amortization", USD_B)),
    ("설비투자", 12, "raw", ("capex", USD_B)),
    ("영업현금흐름", 13, "raw", ("operating_cash_flow", USD_B)),
    ("자기자본", 13, "raw", ("total_equity", USD_B)),
    ("자기자본(전기)", 13, "raw", ("total_equity_prior", USD_B)),
    ("현금", 12, "raw", ("cash", USD_B)),
    ("현금(전기)", 12, "raw", ("cash_prior", USD_B)),
    ("단기차입금", 12, "raw", ("short_term_debt", USD_B)),
    ("단기차입금(전기)", 13, "raw", ("short_term_debt_prior", USD_B)),
    ("장기차입금", 12, "raw", ("long_term_debt", USD_B)),
    ("유동자산", 13, "raw", ("current_assets", USD_B)),
    ("유동자산(전기)", 13, "raw", ("current_assets_prior", USD_B)),
    ("유동부채", 13, "raw", ("current_liabilities", USD_B)),
    ("유동부채(전기)", 13, "raw", ("current_liabilities_prior", USD_B)),
    ("투하자본(전기)", 13, "raw", ("invested_capital_prior", USD_B)),
    ("주식수(백만주)", 13, "raw", ("shares_outstanding", INT)),
]

# Formula columns, appended after the raw block. Each names the raw columns it
# uses by header so the letters stay correct if the layout moves.
YEAR_FORMULAS = [
    ("영업이익(EBIT)", 13, USD_B,
     '=IF(AND(ISNUMBER({세전이익}),ISNUMBER({이자비용})),{세전이익}+{이자비용},'
     'IF(ISNUMBER({영업이익(보고)}),{영업이익(보고)},IF(ISNUMBER({세전이익}),{세전이익},"")))'),
    ("실효세율", 10, PCT,
     '=IF(OR(NOT(ISNUMBER({법인세비용})),NOT(ISNUMBER({세전이익})),{세전이익}<=0),{기본세율},'
     'MIN(MAX({법인세비용}/{세전이익},{세율하한}),{세율상한}))'),
    ("NOPAT", 13, USD_B, '=IF(ISNUMBER({영업이익(EBIT)}),{영업이익(EBIT)}*(1-{실효세율}),"")'),
    ("이자부부채", 13, USD_B,
     '=IF(AND(NOT(ISNUMBER({단기차입금})),NOT(ISNUMBER({장기차입금}))),"",'
     'IF(ISNUMBER({단기차입금}),{단기차입금},0)+IF(ISNUMBER({장기차입금}),{장기차입금},0))'),
    # A bank's balance sheet does not describe invested capital - debt is raw
    # material there, not financing - so these decline to compute for financials
    # rather than emitting a number with no interpretation.
    ("투하자본", 13, USD_B,
     '=IF({업종구분}="금융","",IF(AND(ISNUMBER({자기자본}),ISNUMBER({이자부부채}),ISNUMBER({현금})),'
     '{자기자본}+{이자부부채}-{현금},""))'),
    ("투하자본(평균)", 13, USD_B,
     '=IF({업종구분}="금융","",IF(NOT(ISNUMBER({투하자본})),"",IF(ISNUMBER({투하자본(전기)}),'
     '({투하자본}+{투하자본(전기)})/2,{투하자본})))'),
    ("ROIC", 10, PCT,
     '=IF(AND(ISNUMBER({NOPAT}),ISNUMBER({투하자본(평균)}),{투하자본(평균)}>0),'
     '{NOPAT}/{투하자본(평균)},"")'),
    ("자기자본(평균)", 13, USD_B,
     '=IF(NOT(ISNUMBER({자기자본})),"",IF(ISNUMBER({자기자본(전기)}),'
     '({자기자본}+{자기자본(전기)})/2,{자기자본}))'),
    ("ROE", 10, PCT,
     '=IF(AND(ISNUMBER({순이익}),ISNUMBER({자기자본(평균)}),{자기자본(평균)}<>0),'
     '{순이익}/{자기자본(평균)},"")'),
    ("ROE−ROIC", 11, PCT,
     '=IF(AND(ISNUMBER({ROE}),ISNUMBER({ROIC})),{ROE}-{ROIC},"")'),
    ("EBITDA", 13, USD_B,
     '=IF(AND(ISNUMBER({영업이익(EBIT)}),ISNUMBER({감가상각})),{영업이익(EBIT)}+{감가상각},"")'),
    ("순부채/EBITDA", 12, MULT,
     '=IF(AND(ISNUMBER({이자부부채}),ISNUMBER({현금}),ISNUMBER({EBITDA}),{EBITDA}<>0),'
     '({이자부부채}-{현금})/{EBITDA},"")'),
    ("이자보상배율", 12, MULT,
     '=IF(AND(ISNUMBER({영업이익(EBIT)}),ISNUMBER({이자비용}),{이자비용}<>0),'
     '{영업이익(EBIT)}/{이자비용},"")'),
    ("운전자본(영업)", 13, USD_B,
     '=IF(AND(ISNUMBER({유동자산}),ISNUMBER({유동부채})),'
     '{유동자산}-IF(ISNUMBER({현금}),{현금},0)-({유동부채}-IF(ISNUMBER({단기차입금}),{단기차입금},0)),"")'),
    ("운전자본(전기)", 13, USD_B,
     '=IF(AND(ISNUMBER({유동자산(전기)}),ISNUMBER({유동부채(전기)})),'
     '{유동자산(전기)}-IF(ISNUMBER({현금(전기)}),{현금(전기)},0)'
     '-({유동부채(전기)}-IF(ISNUMBER({단기차입금(전기)}),{단기차입금(전기)},0)),"")'),
    ("운전자본증감", 13, USD_B,
     '=IF(AND(ISNUMBER({운전자본(영업)}),ISNUMBER({운전자본(전기)})),'
     '{운전자본(영업)}-{운전자본(전기)},"")'),
    # The direct definition first; where the classified balance sheet needed for
    # the working-capital swing does not exist, operating cash flow less capex
    # gets there instead - it already embeds D&A and the swing.
    ("주주이익", 13, USD_B,
     '=IF({업종구분}="금융","",'
     'IF(AND(ISNUMBER({순이익}),ISNUMBER({감가상각}),ISNUMBER({설비투자}),ISNUMBER({운전자본증감})),'
     '{순이익}+{감가상각}-{설비투자}-{운전자본증감},'
     'IF(AND(ISNUMBER({영업현금흐름}),ISNUMBER({설비투자})),{영업현금흐름}-{설비투자},"")))'),
    ("주주이익 산출방식", 30, None,
     '=IF({업종구분}="금융","금융업 - 정의되지 않음",'
     'IF(AND(ISNUMBER({순이익}),ISNUMBER({감가상각}),ISNUMBER({설비투자}),ISNUMBER({운전자본증감})),'
     '"순이익+감가상각-설비투자-운전자본증감",'
     'IF(AND(ISNUMBER({영업현금흐름}),ISNUMBER({설비투자})),"영업현금흐름-설비투자","산출불가")))'),
    ("주주이익률", 11, PCT,
     '=IF(AND(ISNUMBER({주주이익}),ISNUMBER({매출}),{매출}<>0),{주주이익}/{매출},"")'),
    ("영업이익률", 11, PCT,
     '=IF(AND(ISNUMBER({영업이익(EBIT)}),ISNUMBER({매출}),{매출}<>0),{영업이익(EBIT)}/{매출},"")'),
]


def sheet_yearly(wb, companies, ref):
    ws = wb.create_sheet("연도별계산")
    r = title_block(
        ws, "연도별 계산 — 원천 수치와 그로부터 나온 모든 비율",
        "파란 셀은 10-K에서 그대로 가져온 원천 수치이고, 검은 셀은 같은 행의 파란 셀로 계산한 "
        "수식입니다. 어떤 비율이든 셀을 클릭하면 수식 입력줄에 계산이 그대로 보입니다. "
        "금액 단위는 십억 USD. 접수번호로 해당 10-K를 특정할 수 있습니다.")

    labels = [c[0] for c in YEAR_COLS] + [f[0] for f in YEAR_FORMULAS]
    widths = [c[1] for c in YEAR_COLS] + [f[1] for f in YEAR_FORMULAS]
    header(ws, r, labels, widths)
    col_of = {name: i for i, name in enumerate(labels, 1)}
    hdr_row = r
    r += 1
    first = r

    for c in companies:
        for y in c["years"]:
            raw = y.get("raw", {})
            for i, (name, _w, kind, spec) in enumerate(YEAR_COLS, 1):
                if kind == "text":
                    v = {"ticker": c["ticker"], "company_name": c["company_name"],
                         "fiscal_year": y.get("fiscal_year"),
                         "period_end": y.get("period_end"),
                         "accession": y.get("accession"),
                         "sector": "금융" if c["sector_treatment"] == "FINANCIAL" else "일반",
                         }[spec]
                    put(ws, r, i, v)
                else:
                    key, fmt = spec
                    v = raw.get(key)
                    if key == "shares_outstanding":
                        v = y.get("shares_outstanding")
                        v = None if v is None else v / 1e6
                    elif v is not None:
                        v = v / 1e9
                    put(ws, r, i, v, fmt, font=BLUE, fill=RAW_FILL)

            for name, _w, fmt, template in YEAR_FORMULAS:
                col = col_of[name]
                f = template
                for label in list(col_of):
                    f = f.replace("{" + label + "}", f"{get_column_letter(col_of[label])}{r}")
                f = (f.replace("{기본세율}", ref["기본 세율"])
                      .replace("{세율하한}", ref["실효세율 하한"])
                      .replace("{세율상한}", ref["실효세율 상한"]))
                font = GREEN if "기준표!" in f else BLACK
                put(ws, r, col, f, fmt, font=font)
            r += 1

    ws.auto_filter.ref = f"A{hdr_row}:{get_column_letter(len(labels))}{r-1}"
    return col_of, first, r - 1


# ===========================================================================
# 품질순위 - the score, with each component's formula on show
# ===========================================================================

SCORE_COLS = [
    ("배점:ROIC", '=IF({ROIC 중앙값}>={t1},25,IF({ROIC 중앙값}>={t2},20,'
                  'IF({ROIC 중앙값}>={t3},15,IF({ROIC 중앙값}>={t4},10,0))))',
     ["ROIC 25점 기준", "ROIC 20점 기준", "ROIC 15점 기준", "ROIC 10점 기준"]),
    ("배점:지속성", '=ROUND(20*{두자릿수 년수}/{관측 년수},0)', []),
    ("배점:신규ROIC", '=IF(NOT(ISNUMBER({신규 ROIC})),0,IF({신규 ROIC}>={t1},20,'
                     'IF({신규 ROIC}>={t2},15,IF({신규 ROIC}>={t3},10,IF({신규 ROIC}>0,5,0)))))',
     ["신규ROIC 20점 기준", "신규ROIC 15점 기준", "신규ROIC 10점 기준"]),
    ("배점:스프레드", '=IF(NOT(ISNUMBER({ROIC−WACC})),0,IF({ROIC−WACC}>={t1},15,'
                    'IF({ROIC−WACC}>={t2},12,IF({ROIC−WACC}>={t3},8,IF({ROIC−WACC}>0,4,0)))))',
     ["스프레드 15점 기준", "스프레드 12점 기준", "스프레드 8점 기준"]),
    ("배점:주주이익", '=IF(NOT(ISNUMBER({주주이익률})),0,IF({주주이익률}>={t1},10,'
                    'IF({주주이익률}>={t2},7,IF({주주이익률}>={t3},4,IF({주주이익률}>0,2,0)))))',
     ["주주이익률 10점 기준", "주주이익률 7점 기준", "주주이익률 4점 기준"]),
    ("배점:순부채", '=IF(NOT(ISNUMBER({순부채/EBITDA})),5,IF({순부채/EBITDA}<={t1},5,'
                  'IF({순부채/EBITDA}<={t2},4,IF({순부채/EBITDA}<={t3},2,0))))',
     ["순부채 5점 기준", "순부채 4점 기준", "순부채 2점 기준"]),
    ("배점:이자보상", '=IF(NOT(ISNUMBER({이자보상배율})),0,IF({이자보상배율}>={t1},5,'
                    'IF({이자보상배율}>={t2},3,IF({이자보상배율}>={t3},1,0))))',
     ["이자보상 5점 기준", "이자보상 3점 기준", "이자보상 1점 기준"]),
]


def sheet_quality(wb, std, ref):
    ws = wb.create_sheet("품질순위")
    r = title_block(
        ws, "품질 순위 — 사업의 질",
        "총점은 오른쪽 7개 배점의 합계이고, 각 배점은 왼쪽 지표와 '기준표'의 문턱값으로 계산한 "
        "수식입니다. 배점 셀을 클릭하면 어떤 조건으로 몇 점이 됐는지 그대로 보입니다. "
        "가격은 여기에 들어가지 않습니다 — 'DCF계산'과 '밸류에이션'에서 따로 다룹니다.")
    labels = ["순위", "티커", "기업", "총점",
              "ROIC 중앙값", "두자릿수 년수", "관측 년수", "지속성 비율",
              "신규 ROIC", "ROIC−WACC", "WACC", "주주이익률", "순부채/EBITDA",
              "이자보상배율", "해자 판정"] + [s[0] for s in SCORE_COLS]
    widths = [6, 8, 26, 8, 12, 12, 11, 11, 12, 11, 9, 12, 13, 12, 38] + [11] * len(SCORE_COLS)
    header(ws, r, labels, widths)
    col = {n: i for i, n in enumerate(labels, 1)}
    r += 1
    first = r

    for i, c in enumerate(std, 1):
        s = c["summary"]
        rev = s.get("_latest_revenue")
        oe = s.get("owner_earnings_normalised")
        put(ws, r, col["순위"], i, INT)
        put(ws, r, col["티커"], c["ticker"])
        put(ws, r, col["기업"], c["company_name"])
        put(ws, r, col["ROIC 중앙값"], s.get("roic_10y_median"), PCT, font=BLUE, fill=RAW_FILL)
        put(ws, r, col["두자릿수 년수"], s.get("roic_above_10pct_years"), INT, font=BLUE, fill=RAW_FILL)
        put(ws, r, col["관측 년수"], s.get("roic_years_observed"), INT, font=BLUE, fill=RAW_FILL)
        put(ws, r, col["지속성 비율"],
            f"={get_column_letter(col['두자릿수 년수'])}{r}/{get_column_letter(col['관측 년수'])}{r}", PCT)
        put(ws, r, col["신규 ROIC"], s.get("incremental_roic"), PCT, font=BLUE, fill=RAW_FILL)
        put(ws, r, col["ROIC−WACC"], s.get("roic_wacc_spread"), PCT, font=BLUE, fill=RAW_FILL)
        put(ws, r, col["WACC"], s.get("wacc"), PCT, font=BLUE, fill=RAW_FILL)
        put(ws, r, col["주주이익률"], (oe / rev) if (oe is not None and rev) else None,
            PCT, font=BLUE, fill=RAW_FILL)
        put(ws, r, col["순부채/EBITDA"], s.get("net_debt_to_ebitda_latest"), MULT,
            font=BLUE, fill=RAW_FILL)
        put(ws, r, col["이자보상배율"], s.get("interest_coverage_latest"), MULT,
            font=BLUE, fill=RAW_FILL)
        put(ws, r, col["해자 판정"], c["_moat"])

        for name, template, thresholds in SCORE_COLS:
            f = template
            for label in col:
                f = f.replace("{" + label + "}", f"{get_column_letter(col[label])}{r}")
            for n, key in enumerate(thresholds, 1):
                f = f.replace("{t%d}" % n, ref[key])
            put(ws, r, col[name], f, INT, font=GREEN if "기준표!" in f else BLACK)

        lo = get_column_letter(col[SCORE_COLS[0][0]])
        hi = get_column_letter(col[SCORE_COLS[-1][0]])
        put(ws, r, col["총점"], f"=SUM({lo}{r}:{hi}{r})", INT, font=BOLD)
        r += 1

    ws.auto_filter.ref = f"A{first-1}:{get_column_letter(len(labels))}{r-1}"
    return first, r - 1


# ===========================================================================
# DCF계산 - the discounting, year by year
# ===========================================================================

def sheet_dcf(wb, valued, ref):
    ws = wb.create_sheet("DCF계산")
    r = title_block(
        ws, "DCF 계산 — 10년치 현금흐름과 할인 과정",
        "회사마다 보수·기준·낙관 3개 행이 있습니다. 1~10년차 열은 각각 그 해 주주이익을 현재가치로 "
        "할인한 값이고, 적정가치는 그 합계에 영구가치의 현재가치를 더한 것입니다. "
        "할인율·영구성장률·성장률 배수는 '기준표'에서 가져옵니다. 금액 단위는 십억 USD.")
    years = 10
    labels = (["티커", "기업", "시나리오", "기준 주주이익", "과거 성장률",
               "적용 성장률", "할인율", "영구성장률"]
              + [f"{i}년차 PV" for i in range(1, years + 1)]
              + ["예측기간 PV 합계", "10년차 현금흐름", "영구가치", "영구가치 PV",
                 "적정가치", "시가총액", "안전마진", "판정"])
    widths = ([9, 22, 10, 14, 12, 12, 10, 12] + [11] * years
              + [15, 15, 13, 14, 13, 13, 12, 12])
    header(ws, r, labels, widths)
    col = {n: i for i, n in enumerate(labels, 1)}
    r += 1
    first = r
    L = get_column_letter
    scen = [("보수", "보수 할인율", "보수 영구성장률", "보수 성장률 배수", "성장률 상한(보수·기준)"),
            ("기준", "기준 할인율", "기준 영구성장률", "기준 성장률 배수", "성장률 상한(보수·기준)"),
            ("낙관", "낙관 할인율", "낙관 영구성장률", "낙관 성장률 배수", "성장률 상한(낙관)")]
    base_row = {}

    for c in valued:
        s = c["summary"]
        for name, rk, tgk, mk, capk in scen:
            put(ws, r, col["티커"], c["ticker"])
            put(ws, r, col["기업"], c["company_name"])
            put(ws, r, col["시나리오"], name)
            put(ws, r, col["기준 주주이익"], bn(s.get("owner_earnings_normalised")),
                USD_B, font=BLUE, fill=RAW_FILL)
            put(ws, r, col["과거 성장률"], s.get("owner_earnings_cagr"), PCT,
                font=BLUE, fill=RAW_FILL)
            g_hist = L(col["과거 성장률"]) + str(r)
            put(ws, r, col["적용 성장률"],
                f'=IF(NOT(ISNUMBER({g_hist})),0.03,MAX(0,MIN({g_hist}*{ref[mk]},{ref[capk]})))',
                PCT, font=GREEN)
            put(ws, r, col["할인율"], f"={ref[rk]}", PCT, font=GREEN)
            put(ws, r, col["영구성장률"], f"={ref[tgk]}", PCT, font=GREEN)

            oe = f"${L(col['기준 주주이익'])}{r}"
            g = f"${L(col['적용 성장률'])}{r}"
            disc = f"${L(col['할인율'])}{r}"
            tg = f"${L(col['영구성장률'])}{r}"
            for t in range(1, years + 1):
                put(ws, r, col[f"{t}년차 PV"],
                    f"={oe}*(1+{g})^{t}/(1+{disc})^{t}", USD_B)
            put(ws, r, col["예측기간 PV 합계"],
                f"=SUM({L(col['1년차 PV'])}{r}:{L(col[f'{years}년차 PV'])}{r})", USD_B, font=BOLD)
            put(ws, r, col["10년차 현금흐름"], f"={oe}*(1+{g})^{years}", USD_B)
            put(ws, r, col["영구가치"],
                f"=IF({disc}<={tg},\"\",{L(col['10년차 현금흐름'])}{r}*(1+{tg})/({disc}-{tg}))", USD_B)
            put(ws, r, col["영구가치 PV"],
                f"=IF(ISNUMBER({L(col['영구가치'])}{r}),{L(col['영구가치'])}{r}/(1+{disc})^{years},\"\")",
                USD_B)
            put(ws, r, col["적정가치"],
                f"=IF(ISNUMBER({L(col['영구가치 PV'])}{r}),"
                f"{L(col['예측기간 PV 합계'])}{r}+{L(col['영구가치 PV'])}{r},\"\")", USD_B, font=BOLD)
            put(ws, r, col["시가총액"], bn(c.get("market_cap_usd")), USD_B, font=BLUE, fill=RAW_FILL)
            iv, mc = f"{L(col['적정가치'])}{r}", f"{L(col['시가총액'])}{r}"
            put(ws, r, col["안전마진"],
                f'=IF(AND(ISNUMBER({iv}),{iv}<>0,ISNUMBER({mc})),({iv}-{mc})/{iv},"")', PCT)
            mos = f"{L(col['안전마진'])}{r}"
            put(ws, r, col["판정"],
                f'=IF(NOT(ISNUMBER({mos})),"판정불가",IF({mos}>={ref["안전마진 투자가능 기준"]},'
                f'"투자가능",IF({mos}>=0,"경계선","범위밖")))', font=GREEN)
            if name == "기준":
                base_row[c["ticker"]] = r
            r += 1

    ws.auto_filter.ref = f"A{first-1}:{L(len(labels))}{r-1}"
    return base_row, col, first, r - 1


# ===========================================================================

def sheet_valuation(wb, valued, base_row, dcf_col, ref):
    ws = wb.create_sheet("밸류에이션")
    r = title_block(
        ws, "밸류에이션 요약 — 원칙 5·6",
        "적정가치와 안전마진은 'DCF계산' 시트의 기준 시나리오 행을 그대로 참조합니다(초록 글씨). "
        "셀을 클릭하면 어느 행에서 왔는지 보이고, 그 행에서 10년치 할인 과정을 볼 수 있습니다.")
    labels = ["티커", "기업", "시가총액", "정규화 주주이익", "최근 순이익", "최근 자기자본",
              "최근 매출", "PER", "PBR", "PSR", "기준 적정가치", "P/적정가치", "안전마진",
              "판정", "정규화 방식"]
    header(ws, r, labels, [9, 26, 14, 15, 14, 14, 14, 9, 9, 9, 14, 12, 12, 12, 50])
    col = {n: i for i, n in enumerate(labels, 1)}
    r += 1
    first = r
    L = get_column_letter
    for c in valued:
        s = c["summary"]
        latest = c["years"][-1] if c["years"] else {}
        dr = base_row[c["ticker"]]
        put(ws, r, col["티커"], c["ticker"])
        put(ws, r, col["기업"], c["company_name"])
        put(ws, r, col["시가총액"], bn(c.get("market_cap_usd")), USD_B, font=BLUE, fill=RAW_FILL)
        put(ws, r, col["정규화 주주이익"], bn(s.get("owner_earnings_normalised")), USD_B,
            font=BLUE, fill=RAW_FILL)
        put(ws, r, col["최근 순이익"], bn(latest.get("net_income")), USD_B, font=BLUE, fill=RAW_FILL)
        put(ws, r, col["최근 자기자본"], bn(latest.get("total_equity")), USD_B, font=BLUE, fill=RAW_FILL)
        put(ws, r, col["최근 매출"], bn(latest.get("revenue")), USD_B, font=BLUE, fill=RAW_FILL)
        mc = f"{L(col['시가총액'])}{r}"
        put(ws, r, col["PER"],
            f'=IF(AND(ISNUMBER({L(col["최근 순이익"])}{r}),{L(col["최근 순이익"])}{r}>0),'
            f'{mc}/{L(col["최근 순이익"])}{r},"")', MULT)
        put(ws, r, col["PBR"],
            f'=IF(AND(ISNUMBER({L(col["최근 자기자본"])}{r}),{L(col["최근 자기자본"])}{r}<>0),'
            f'{mc}/{L(col["최근 자기자본"])}{r},"")', MULT)
        put(ws, r, col["PSR"],
            f'=IF(AND(ISNUMBER({L(col["최근 매출"])}{r}),{L(col["최근 매출"])}{r}<>0),'
            f'{mc}/{L(col["최근 매출"])}{r},"")', MULT)
        put(ws, r, col["기준 적정가치"],
            f"=DCF계산!{L(dcf_col['적정가치'])}{dr}", USD_B, font=GREEN)
        iv = f"{L(col['기준 적정가치'])}{r}"
        put(ws, r, col["P/적정가치"],
            f'=IF(AND(ISNUMBER({iv}),{iv}<>0),{mc}/{iv},"")', MULT)
        put(ws, r, col["안전마진"], f"=DCF계산!{L(dcf_col['안전마진'])}{dr}", PCT, font=GREEN)
        put(ws, r, col["판정"], f"=DCF계산!{L(dcf_col['판정'])}{dr}", font=GREEN)
        put(ws, r, col["정규화 방식"], s.get("owner_earnings_normalisation"), font=NOTE)
        r += 1
    ws.auto_filter.ref = f"A{first-1}:{L(len(labels))}{r-1}"


def sheet_financials(wb, fin):
    ws = wb.create_sheet("금융9개사")
    r = title_block(ws, "금융업 — 별도 기준",
                    "ROIC와 주주이익은 산출하지 않았습니다. 은행·보험의 대차대조표에서 부채는 "
                    "영업 자산이고 유동자산·유동부채 구분이 없어 투하자본과 운전자본이 정의되지 "
                    "않기 때문입니다. PER·PBR은 이 시트 안의 원천 수치로 계산한 수식입니다.")
    labels = ["티커", "기업", "SIC", "업종", "ROE 중앙값", "최근 ROE",
              "시가총액", "최근 순이익", "최근 자기자본", "PER", "PBR"]
    header(ws, r, labels, [9, 26, 7, 30, 12, 11, 14, 14, 14, 9, 9])
    col = {n: i for i, n in enumerate(labels, 1)}
    r += 1
    first = r
    L = get_column_letter
    for c in fin:
        s = c["summary"]
        latest = c["years"][-1] if c["years"] else {}
        put(ws, r, col["티커"], c["ticker"])
        put(ws, r, col["기업"], c["company_name"])
        put(ws, r, col["SIC"], c.get("sic"))
        put(ws, r, col["업종"], c.get("sic_description"))
        put(ws, r, col["ROE 중앙값"], s.get("roe_10y_median"), PCT, font=BLUE, fill=RAW_FILL)
        put(ws, r, col["최근 ROE"], s.get("roe_latest"), PCT, font=BLUE, fill=RAW_FILL)
        put(ws, r, col["시가총액"], bn(c.get("market_cap_usd")), USD_B, font=BLUE, fill=RAW_FILL)
        put(ws, r, col["최근 순이익"], bn(latest.get("net_income")), USD_B, font=BLUE, fill=RAW_FILL)
        put(ws, r, col["최근 자기자본"], bn(latest.get("total_equity")), USD_B, font=BLUE, fill=RAW_FILL)
        mc = f"{L(col['시가총액'])}{r}"
        put(ws, r, col["PER"],
            f'=IF(AND(ISNUMBER({L(col["최근 순이익"])}{r}),{L(col["최근 순이익"])}{r}>0),'
            f'{mc}/{L(col["최근 순이익"])}{r},"")', MULT)
        put(ws, r, col["PBR"],
            f'=IF(AND(ISNUMBER({L(col["최근 자기자본"])}{r}),{L(col["최근 자기자본"])}{r}<>0),'
            f'{mc}/{L(col["최근 자기자본"])}{r},"")', MULT)
        r += 1
    ws.auto_filter.ref = f"A{first-1}:{L(len(labels))}{r-1}"


def sheet_excluded(wb, std, fin):
    ws = wb.create_sheet("밸류에이션제외")
    r = title_block(ws, "밸류에이션을 내지 않은 기업과 그 이유",
                    "빈칸을 추정으로 메우지 않고 이유를 적었습니다.")
    header(ws, r, ["티커", "기업", "구분", "사유"], [9, 26, 20, 96])
    r += 1
    for c in std:
        s = c["summary"]
        if s["valuation_status"] == "NEGATIVE_OWNER_EARNINGS":
            put(ws, r, 1, c["ticker"])
            put(ws, r, 2, c["company_name"])
            put(ws, r, 3, "주주이익 음수")
            put(ws, r, 4, f"정규화 주주이익 {s['owner_earnings_normalised']/1e9:.1f}십억 달러. "
                          f"설비투자와 운전자본이 순이익을 넘어서 할인할 현금흐름이 없습니다. "
                          f"데이터 공백이 아니라 사업에 대한 판정입니다.")
            r += 1
    for c in fin:
        put(ws, r, 1, c["ticker"])
        put(ws, r, 2, c["company_name"])
        put(ws, r, 3, "금융업 (SIC 6000~6799)")
        put(ws, r, 4, "주주이익이 정의되지 않습니다. 설비투자와 운전자본은 은행이 자본을 "
                      "어떻게 굴리는지 설명하지 못합니다. '금융9개사' 시트에서 ROE 기준으로 평가했습니다.")
        r += 1


def sheet_verification(wb, analysis, verification):
    ws = wb.create_sheet("검증결과")
    r = title_block(ws, "PHASE 6 — 출처 재검증",
                    "각 수치를 해당 10-K 원문의 인라인 XBRL과 대조했습니다. "
                    "companyfacts API가 아니라 제출서류 자체를 다시 읽어 비교한 결과입니다.")
    header(ws, r, ["티커", "판정", "항목 통과", "접수번호 확인", "주식수 재현", "주가 재조회",
                   "검증한 제출서류", "companyfacts 출처"], [9, 8, 11, 13, 12, 12, 72, 72])
    r += 1
    first = r
    src = {c["ticker"]: c.get("source_url") for c in analysis["companies"]}
    for t, v in verification["companies"].items():
        n_pass = sum(1 for m in v["metrics"] if m["result"] == "PASS")
        n_ck = sum(1 for m in v["metrics"] if m["result"] in ("PASS", "FAIL"))
        for i, val in enumerate([t, v["verdict"], f"{n_pass}/{n_ck}", v["accessions"]["result"],
                                 v["cover_shares"]["result"], v["price"]["result"],
                                 v.get("filing_verified"), src.get(t)], 1):
            put(ws, r, i, val)
        r += 1
    ws.auto_filter.ref = f"A{first-1}:H{r-1}"
    r += 1
    for label, node in (("무위험수익률", verification["macro"]["risk_free_rate"]),
                        ("주식위험프리미엄", verification["macro"]["equity_risk_premium"])):
        put(ws, r, 1, label, font=BOLD)
        put(ws, r, 2, node["result"])
        put(ws, r, 3, node["source_url"])
        r += 1


def sheet_universe(wb, universe, analysis):
    ws = wb.create_sheet("유니버스")
    r = title_block(ws, "유니버스 52개사 — 선정 근거",
                    "미국 50개사는 시가총액 상위 기준으로 선정했습니다. 선정에 쓴 외부 시가총액은 "
                    "구성 결정에만 사용했고 보고서의 어떤 수치로도 인용하지 않았습니다.")
    header(ws, r, ["티커", "기업", "국가", "분석 방식", "CIK", "선정 순위",
                   "선정용 시가총액", "계산된 시가총액", "주식수 출처", "비고"],
           [10, 30, 7, 16, 11, 10, 16, 16, 50, 60])
    r += 1
    first = r
    rec = {c["ticker"]: c for c in analysis["companies"]}
    for c in universe["companies"]:
        a = rec.get(c["ticker"], {})
        sel = c.get("selection") or {}
        notes = []
        if c.get("cik_override"):
            notes.append(f"CIK 교정: {c['cik_override']['reason']}")
        if c.get("reason"):
            notes.append(c["reason"])
        cov = a.get("shares_cover_page") or {}
        if cov.get("incomplete"):
            notes.append(cov.get("conversion_note", ""))
        vals = [c["ticker"], c["company_name"], c["exchange_country"],
                "정량" if c.get("analysis_mode") == "QUANTITATIVE" else "정성만",
                c.get("cik"), sel.get("screen_rank"), bn(sel.get("screen_market_cap_usd")),
                bn(a.get("market_cap_usd")), a.get("shares_source"), " / ".join(notes)]
        fmts = [None, None, None, None, None, INT, USD_B, USD_B, None, None]
        for i, (v, f) in enumerate(zip(vals, fmts), 1):
            put(ws, r, i, v, f)
        r += 1
    ws.auto_filter.ref = f"A{first-1}:J{r-1}"
    r += 1
    put(ws, r, 1, "제외: SPCX(SpaceX) — 비상장으로 10-K 제출 이력이 없어 감사받은 연간 "
                  "시계열이 존재하지 않습니다. 차순위인 IBM을 승격했습니다.", font=NOTE)


# ===========================================================================
# 산식정의 - the data dictionary
# ===========================================================================

DEFINITIONS = [
    ("연도별계산", "매출 / 세전이익 / 이자비용 / 법인세비용 / 순이익 / 감가상각 / 설비투자 / "
     "자기자본 / 현금 / 단기·장기차입금 / 유동자산 / 유동부채",
     "원천 (파란 셀)", "각 회계연도 10-K의 XBRL 태그 값을 그대로 가져온 것. 가공하지 않았습니다.",
     "SEC EDGAR companyfacts, 접수번호 열로 해당 10-K 특정 가능"),
    ("연도별계산", "영업이익(EBIT)", "세전이익 + 이자비용",
     "세전이익과 이자비용이 모두 있으면 둘을 더합니다. 릴리·엑슨모빌·IBM·머크·셰브론은 영업이익 "
     "소계를 아예 태깅하지 않아, 태그가 있는 회사만 보고된 영업이익을 쓰면 ROIC가 회사마다 다른 "
     "것을 뜻하게 됩니다. 그래서 전 종목을 같은 방식으로 맞췄고, 이자비용이 없는 회사만 보고된 "
     "영업이익으로 대체합니다.", "같은 행의 원천 셀"),
    ("연도별계산", "실효세율", "법인세비용 ÷ 세전이익 (5~40%로 제한)",
     "세전이익이 없거나 음수면 기본 세율 21%를 씁니다. 범위를 벗어나는 값은 일회성 항목의 결과라 "
     "한계 세율을 대표하지 못하므로 상·하한을 둡니다.", "기준표의 세율 셀 3개"),
    ("연도별계산", "NOPAT", "영업이익(EBIT) × (1 − 실효세율)",
     "세후 영업이익. ROIC의 분자입니다.", "같은 행"),
    ("연도별계산", "이자부부채", "단기차입금 + 장기차입금",
     "한쪽만 태깅된 경우 나머지는 0으로 봅니다. 잔액이 0이면 태그를 생략하는 것이 관행이기 "
     "때문입니다. 양쪽 다 없으면 공란으로 둡니다.", "같은 행"),
    ("연도별계산", "투하자본", "자기자본 + 이자부부채 − 현금",
     "현금을 빼는 이유는 놀고 있는 현금이 사업에 투입된 자본이 아니기 때문입니다. 빼지 않으면 "
     "현금이 많은 대형주의 ROIC가 실제보다 낮게 나옵니다.", "같은 행"),
    ("연도별계산", "투하자본(평균)", "(당기 투하자본 + 전기 투하자본) ÷ 2",
     "기말 잔고를 쓰면 연말에 자본을 조달한 회사가 한 해 내내 그 자본으로 벌어들인 것처럼 "
     "보입니다. 전기 값이 없으면 당기 값을 그대로 씁니다.", "같은 행"),
    ("연도별계산", "ROIC", "NOPAT ÷ 투하자본(평균)", "원칙 1의 핵심 지표.", "같은 행"),
    ("연도별계산", "ROE", "순이익 ÷ 자기자본(평균)", "원칙 2에서 ROIC와 비교합니다.", "같은 행"),
    ("연도별계산", "ROE−ROIC", "ROE − ROIC",
     "차이가 크면 ROE가 사업의 수익성이 아니라 레버리지에서 나온다는 신호입니다. 다만 자사주 "
     "매입으로 자기자본이 줄어도 같은 형태로 벌어지므로 부채 지표를 함께 봐야 합니다.", "같은 행"),
    ("연도별계산", "EBITDA", "영업이익(EBIT) + 감가상각", "", "같은 행"),
    ("연도별계산", "순부채/EBITDA", "(이자부부채 − 현금) ÷ EBITDA", "레버리지 건전성.", "같은 행"),
    ("연도별계산", "이자보상배율", "영업이익(EBIT) ÷ 이자비용", "이자를 몇 배로 감당하는지.", "같은 행"),
    ("연도별계산", "운전자본(영업)", "유동자산 − 현금 − (유동부채 − 단기차입금)",
     "현금과 단기차입금을 빼는 것이 핵심입니다. 유동자산에는 현금이 포함되므로 그대로 두면 "
     "현금을 쌓은 회사가 운전자본을 소모한 것처럼 보입니다. 단기차입금은 영업이 아니라 "
     "자금조달이라 같은 이유로 뺍니다.", "같은 행"),
    ("연도별계산", "운전자본증감", "당기 운전자본(영업) − 전기 운전자본(영업)",
     "양수면 그해 운전자본이 현금을 흡수했다는 뜻입니다.", "같은 행"),
    ("연도별계산", "주주이익", "순이익 + 감가상각 − 설비투자 − 운전자본증감",
     "버핏의 owner earnings 정의입니다. 네 항목이 모두 같은 회계기간이어야 합니다.", "같은 행"),
    ("연도별계산", "주주이익률 / 영업이익률", "각각 주주이익 ÷ 매출, 영업이익(EBIT) ÷ 매출", "", "같은 행"),
    ("품질순위", "ROIC 중앙값 · 신규 ROIC · ROIC−WACC · WACC 등 (파란 셀)",
     "여러 해에 걸친 집계값", "연도별계산의 여러 행을 걸쳐 계산한 값이라 한 행의 수식으로는 "
     "표현되지 않습니다. ROIC 중앙값은 관측된 모든 해의 ROIC 중앙값이고, 신규 ROIC는 "
     "(기간말 NOPAT − 기간초 NOPAT) ÷ (기간말 투하자본 − 기간초 투하자본)입니다.",
     "연도별계산 시트의 해당 기업 행들"),
    ("품질순위", "지속성 비율", "두자릿수 년수 ÷ 관측 년수",
     "ROIC가 10%를 넘긴 해의 비율. 높은 평균보다 꾸준함이 해자에 가깝다는 것이 원칙 1의 취지입니다.",
     "같은 행"),
    ("품질순위", "배점 7개 열", "각 지표를 기준표의 문턱값과 비교한 조건식",
     "셀을 클릭하면 어떤 조건에서 몇 점이 나왔는지 그대로 보입니다. 배점: ROIC 25 · 지속성 20 · "
     "신규 ROIC 20 · 스프레드 15 · 주주이익률 10 · 순부채 5 · 이자보상 5 (합계 100).",
     "같은 행 + 기준표"),
    ("품질순위", "총점", "배점 7개 열의 합계", "백분위가 아니라 절대 기준입니다.", "같은 행"),
    ("DCF계산", "적용 성장률", "과거 주주이익 성장률 × 시나리오 배수, 상한 적용",
     "과거 성장률이 없으면 3%를 씁니다. 상한은 보수·기준 10%, 낙관 15%입니다 — 10년을 그보다 "
     "빠르게 복리성장한다고 보지 않습니다.", "기준표"),
    ("DCF계산", "1~10년차 PV", "기준 주주이익 × (1+성장률)^t ÷ (1+할인율)^t",
     "t년차 주주이익을 현재가치로 할인한 값.", "같은 행"),
    ("DCF계산", "영구가치", "10년차 현금흐름 × (1+영구성장률) ÷ (할인율 − 영구성장률)",
     "고든 성장 모형. 할인율이 영구성장률 이하면 값이 발산하므로 공란으로 둡니다.", "같은 행"),
    ("DCF계산", "적정가치", "예측기간 PV 합계 + 영구가치 PV", "", "같은 행"),
    ("DCF계산", "안전마진", "(적정가치 − 시가총액) ÷ 적정가치",
     "양수면 적정가치보다 싸게 거래된다는 뜻입니다.", "같은 행"),
    ("DCF계산", "판정", "안전마진 ≥ 30% 투자가능 / ≥ 0% 경계선 / 그 외 범위밖",
     "목표주가가 아니라 '지금 투자 가능 구간 안인가'에 대한 답입니다.", "기준표"),
    ("밸류에이션", "PER / PBR / PSR", "시가총액 ÷ (최근 순이익 / 자기자본 / 매출)",
     "PER은 순이익이 양수일 때만 계산합니다. 적자에서는 의미가 없습니다.", "같은 행"),
    ("밸류에이션", "기준 적정가치 · 안전마진 · 판정", "DCF계산 시트의 기준 시나리오 행 참조",
     "초록 글씨는 다른 시트를 가리킵니다. 클릭하면 해당 행으로 따라갈 수 있습니다.", "DCF계산"),
    ("전체", "시가총액", "10-K 표지의 주식수 × 최근 종가",
     "표지 주식수는 결산일보다 최신이고 클래스별로 나뉘어 있어 합산했습니다. 버크셔는 A주를 "
     "B주 1,500배로 환산했습니다. 비자는 클래스 B·C의 전환비율이 이사회 재량이라 클래스 A만 "
     "썼고, 그만큼 시가총액이 과소평가돼 있습니다.",
     "SEC 10-K 표지 인라인 XBRL + 거래소 종가"),
    ("전체", "주식수 시계열", "액면분할 기준을 통일한 뒤 비교",
     "각 연도 주식수는 그해 10-K의 단위로 기록되는데 10-K는 직전 2년까지만 소급 수정하므로, "
     "분할을 가로지르면 신주 발행처럼 보입니다. 조정 전 엔비디아는 연 41% 희석으로 나왔으나 "
     "실제로는 4:1과 10:1 분할이었습니다. 정수가 아닌 비율(델 1806:1000, GE 1281:1000)은 "
     "스핀오프에 따른 주가 조정이라 적용하지 않았습니다.", "거래소 분할 이력"),
]


def sheet_definitions(wb):
    ws = wb.create_sheet("산식정의")
    r = title_block(ws, "산식 정의 — 각 열이 어떻게 계산되는가",
                    "시트별로 열 하나하나의 정의와 산식, 그리고 왜 그렇게 계산했는지를 적었습니다.")
    header(ws, r, ["시트", "열", "산식", "설명", "참조 대상"], [14, 34, 40, 86, 34])
    r += 1
    for sheet, colname, formula, desc, source in DEFINITIONS:
        put(ws, r, 1, sheet, font=BOLD)
        put(ws, r, 2, colname)
        put(ws, r, 3, formula)
        c = put(ws, r, 4, desc)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        put(ws, r, 5, source, font=NOTE)
        ws.row_dimensions[r].height = max(15, 12 * (len(desc) // 55 + 1))
        r += 1


def sheet_summary(wb, analysis, verification, std, fin):
    ws = wb.create_sheet("요약")
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 92
    r = title_block(ws, "버핏 기준 52개 기업 분석",
                    f"생성 {analysis['generated_at_utc'][:16].replace('T', ' ')} UTC · "
                    f"미국 50개사 정량 + 한국 2개사 정성")

    valued = [c for c in std if c["summary"]["valuation_status"] == "VALUED"]
    investable = [c for c in valued if c["summary"]["dcf"]["base"]["verdict"] == "INVESTABLE_RANGE"]
    borderline = [c for c in valued if c["summary"]["dcf"]["base"]["verdict"] == "BORDERLINE"]
    negoe = [c for c in std if c["summary"]["valuation_status"] == "NEGATIVE_OWNER_EARNINGS"]
    above = sum(1 for c in std if (c["summary"].get("roic_wacc_spread") or -1) > 0)
    v = verification["summary"]

    blocks = [
        ("이 파일 읽는 법", None),
        ("글자 색", "파란 글씨는 10-K나 시세에서 그대로 가져온 원천 수치입니다. 검은 글씨는 같은 "
                  "시트에서 계산한 수식이고, 초록 글씨는 다른 시트를 참조하는 수식입니다. "
                  "노란 셀은 바꿔볼 수 있는 가정입니다."),
        ("계산을 따라가는 순서", "연도별계산 → 품질순위 → DCF계산 → 밸류에이션 순으로 보시면 "
                          "원천 수치가 비율이 되고, 비율이 점수와 적정가치가 되는 과정이 "
                          "이어집니다. 각 열의 정의와 산식은 '산식정의' 시트에 정리했습니다."),
        ("가정을 바꾸려면", "'기준표' 시트의 노란 셀을 바꾸면 배점과 적정가치가 함께 움직입니다. "
                      "예를 들어 기준 할인율을 10%에서 9%로 낮추면 모든 적정가치가 올라갑니다."),
        ("", None),
        ("핵심 결론", f"비금융 {len(std)}개사 중 {above}개사가 ROIC로 자본비용을 넘겼지만, "
                  f"할인율 10% 기준 DCF에서 안전마진 30% 이상인 기업은 {len(investable)}개입니다."),
        ("경계선", f"{len(borderline)}개사 ({', '.join(c['ticker'] for c in borderline) or '없음'}). "
                f"나머지는 전부 적정가치 추정 범위 밖에서 거래되고 있습니다."),
        ("주주이익 음수", f"{len(negoe)}개사 ({', '.join(c['ticker'] for c in negoe)}). 이익이 "
                    f"없어서가 아니라 설비투자와 운전자본이 순이익을 넘어서기 때문이며, "
                    f"데이터 공백이 아니라 사업 판정입니다."),
        ("금융업 처리", f"{len(fin)}개사는 ROIC·주주이익을 산출하지 않았습니다. 은행에서 부채는 "
                   f"자금조달이 아니라 원재료라 투하자본이 정의되지 않습니다. SIC 6000~6799 기준."),
        ("한국 2개사", "삼성전자·SK하이닉스는 정성 분석만 했습니다. DART 접근에 인증키가 필요해 "
                  "감사받은 1차 출처를 확보할 수 없었고, 기억으로 숫자를 채우지 않았습니다."),
        ("검증 (PHASE 6)", f"{v['companies_passed']}/{v['companies_checked']}개사 통과. 각 수치를 "
                       f"해당 10-K 원문의 인라인 XBRL과 대조하고, 인용 접수번호가 EDGAR에 "
                       f"실재하는지 확인하고, 주식수를 재파싱으로 재현했습니다."),
        ("", None),
        ("주의", "이 분석은 '지금 투자 가능 구간 안인가'라는 질문에만 답합니다. 목표주가가 아니며, "
               "DCF는 3개 시나리오의 범위로만 읽어야 합니다. 비자의 시가총액은 클래스 B·C "
               "전환분이 빠져 과소평가돼 있어 실제 안전마진은 표시된 값보다 나쁩니다."),
    ]
    for label, text in blocks:
        if not label:
            r += 1
            continue
        a = put(ws, r, 1, label, font=BOLD)
        a.alignment = Alignment(vertical="top")
        if text is None:
            a.font = TITLE_FONT
            r += 1
            continue
        b = put(ws, r, 2, text, font=BLACK)
        b.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = max(15, 13 * (len(text) // 62 + 1))
        r += 1


def main():
    with open(os.path.join(WORK, "analysis.json")) as f:
        analysis = json.load(f)
    with open(os.path.join(WORK, "verification.json")) as f:
        verification = json.load(f)
    with open(os.path.join(WORK, "universe.json")) as f:
        universe = json.load(f)

    sys.path.insert(0, WORK)
    import report as rp

    for c in analysis["companies"]:
        c["summary"]["_latest_revenue"] = next(
            (r["revenue"] for r in reversed(c["years"]) if r.get("revenue")), None)
        c["_score"], c["_detail"] = rp.quality_score(c["summary"])
        c["_moat"] = rp.moat_read(c["summary"])

    std = [c for c in analysis["companies"] if c["sector_treatment"] == "STANDARD"]
    fin = [c for c in analysis["companies"] if c["sector_treatment"] == "FINANCIAL"]
    std.sort(key=lambda c: (-c["_score"], -(c["summary"].get("roic_10y_median") or -9)))
    fin.sort(key=lambda c: -(c["summary"].get("roe_10y_median") or -9))
    valued = [c for c in std if c["summary"]["valuation_status"] == "VALUED"]
    valued.sort(key=lambda c: -(c["summary"]["dcf"]["base"]["margin_of_safety"] or -99))

    wb = Workbook()
    wb.remove(wb.active)
    sheet_summary(wb, analysis, verification, std, fin)
    sheet_definitions(wb)
    ref = sheet_assumptions(wb, analysis)
    sheet_yearly(wb, analysis["companies"], ref)
    sheet_quality(wb, std, ref)
    base_row, dcf_col, _, _ = sheet_dcf(wb, valued, ref)
    sheet_valuation(wb, valued, base_row, dcf_col, ref)
    sheet_financials(wb, fin)
    sheet_excluded(wb, std, fin)
    sheet_verification(wb, analysis, verification)
    sheet_universe(wb, universe, analysis)

    # openpyxl writes formulas with no cached result, so a reader that trusts
    # the cached value sees blanks. This asks the application to recalculate the
    # whole book on open, which Excel, LibreOffice and Sheets all honour.
    wb.calculation = CalcProperties(fullCalcOnLoad=True)
    wb.save(OUT)
    print(f"xlsx -> {OUT}")
    print(f"sheets: {wb.sheetnames}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
